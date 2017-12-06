# create a dictionary of rules like {"ch":'щ'}
# presume data stored in a excel-'file' with known 'sheetname'


def get_rules(file="./rules.xlsx", sheetname='rules', debug="N"):
    """fill dictionary with rules and return it
    \n  file -
    \n  sheetname -
    \n  debug - option to print debug messages to console (in case debug = 'Y')
    """
    deb_print(debug, "Welcome to 'get_rules()' function")
    result = {}
    # open excel-'file'
    # use excel library 'openpyxl'
    import openpyxl
    try:
        deb_print(debug, 'Opening a file ', file)
        wb = openpyxl.load_workbook(file)
        # deb_print(debug, 'List of sheets: ', wb.get_sheet_names())
    except Exception as wb_ex:
        deb_print(debug, wb_ex)
    else:

        try:
            obj_rules_sheet = wb.get_sheet_by_name(sheetname)
        except Exception as ws_ex:
            deb_print(debug, ws_ex)
        else:
            deb_print(debug, "Found '", obj_rules_sheet.title, "' sheet")
            deb_print(debug, "Adding rules..")
            i = 1
            while obj_rules_sheet.cell(row=i, column=1).value is not None:
                result[obj_rules_sheet.cell(row=i, column=1).value] = \
                    obj_rules_sheet.cell(row=i, column=2).value
                i += 1
        deb_print(debug, "Closing the file...")
        wb.close()
    finally:
        deb_print(debug, "Return result. \nEnd of function.\n\n")
        return result


def deb_print(debug, *some_strings):
    """
    prints while debugging
    :param debug:
    :param some_strings:
    :return:
    """
    if debug == "Y":
        print(*some_strings)


if __name__ == '__main__':
    rules = get_rules(debug='N')
    print(rules)
